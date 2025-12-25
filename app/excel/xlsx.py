import re
import pandas as pd
from io import StringIO
from PIL import Image
from io import BytesIO
from typing import List, Dict, Optional
from openpyxl import load_workbook
from html import escape


def read_xlsx_cols(
    filepath: str,
    columns: Optional[List[str]] = None,
) -> Dict[str, List[Dict[str, str]]]:
    """
    从 Excel 文件中提取每个 sheet 的指定列内容，返回的字典包含每个 sheet
    :param filepath: Excel 文件路径（支持 .xls 或 .xlsx）
    :param columns: 需要提取的列名列表，如 ["题目","试题","选项个数","答案","解析","归属"]
                    如果为 None，则提取所有列
    :return: dict，key = sheet 名称, value = 该 sheet 的记录列表
    """
    xls = pd.ExcelFile(filepath)
    sheet_data = {}
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)
        if columns:
            available_columns = [col for col in columns if col in df.columns]
            df = df[available_columns]
        records = df.fillna("").to_dict(orient="records")
        sheet_data[sheet_name] = records

    return sheet_data


def get_images_map(sheet):
    """获取 sheet 内所有图片的"""
    img_map = {}
    max_row = 0
    max_col = 0
    for img in getattr(sheet, "_images", []):
        if hasattr(img.anchor, "_from"):  # 单元格锚点
            row = img.anchor._from.row + 1  # 从 0 开始，要 +1
            col = img.anchor._from.col + 1
            max_row = max(max_row, row)
            max_col = max(max_col, col)
            img_map[(row, col)] = img._data()
    return max_row, max_col, img_map


def get_image_type(data):
    """获取sheet内图片的类型"""
    with Image.open(BytesIO(data)) as img:
        return img.format.lower()  # 'JPEG', 'PNG' → 转小写


def is_empty_row(cols):
    """判断行中所有列是否全为空"""
    for cell in cols:
        if (cell.text or "").strip() != "":
            return False
        # 或者有结构属性（即使无文本，多行合并也可能有意义）
        if cell.get("rowspan"):
            return False
    return True


def text_process(value_str):
    """尝试转为数值并做有效位数处理,单元格内容的换行用标签代替"""
    if value_str is None:
        return ""
    value_str = escape(value_str)
    try:
        num = float(value_str)
        if num.is_integer():
            return int(num)
        else:
            return round(num, 4)
    except (ValueError, TypeError):
        # 使用换行标签保证内容为一行
        return re.sub(r"\r\n|\r|\n", "<br>", value_str)


def sheet_to_html(sheet_name, sheet):
    """高性能转换 Excel 表格为 HTML，支持合并单元格"""
    html = []
    html.append(f"<table><caption>{sheet_name}</caption>")

    # ================= 预处理合并区域 =================
    merged_ranges = list(sheet.merged_cells.ranges)
    merge_start_map = {}  # 构建合并起点映射: (min_row, min_col) -> (rowspan, colspan)
    merged_covered_set = set()  # 构建被覆盖单元格集合
    for mr in merged_ranges:
        min_r, min_c, max_r, max_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        rowspan = max_r - min_r + 1
        colspan = max_c - min_c + 1
        merge_start_map[(min_r, min_c)] = (rowspan, colspan)

        # 记录所有被覆盖的非起点单元格
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if (r, c) != (min_r, min_c):
                    merged_covered_set.add((r, c))

    # 获取图片映射及其单元格锚点
    max_img_row, max_img_col, img_map = get_images_map(sheet)  # 保持原逻辑

    # 结合图像锚点单元格计算最大的行列
    max_row = max(sheet.max_row, max_img_row)
    max_col = max(sheet.max_column, max_img_col)

    # 计算全表的起始列号
    start_col = 999999  # 表格开头空白的列
    for row_idx, row in enumerate(  # 使用 iter_rows 批量读取，提升性能
        sheet.iter_rows(
            min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False
        ),
        1,
    ):
        cells = list(row)  # 这行所有单元格对象
        for col_idx, cell in enumerate(cells, 1):
            # 检查是否是合并起点（用预构建的 map）
            is_merged_start = (row_idx, col_idx) in merge_start_map
            # 检查是否有值
            has_value = cell.value not in (None, "")
            if has_value or is_merged_start:
                if col_idx < start_col:
                    start_col = col_idx

    for row_idx, row in enumerate(  # 使用 iter_rows 批量读取，提升性能
        sheet.iter_rows(
            min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False
        ),
        1,
    ):
        cells = list(row)  # 这行所有单元格对象
        # 找到该行最小/最大有效列（有值 或 是合并起点）
        min_col = None
        max_col_in_row = 0
        for col_idx, cell in enumerate(cells, 1):
            # 检查是否是合并起点（用预构建的 map）
            is_merged_start = (row_idx, col_idx) in merge_start_map
            # 检查是否有值
            has_value = cell.value not in (None, "")
            # 检查是否有图片
            has_img = (row_idx, col_idx) in img_map
            if has_value or is_merged_start or has_img:
                if min_col is None:
                    min_col = col_idx
                if col_idx > max_col_in_row:
                    max_col_in_row = col_idx

        # 整行空，跳过
        if min_col is None:
            continue

        html.append("<tr>")
        col_idx = start_col
        while col_idx <= max_col_in_row:
            # 跳过被合并覆盖的单元格
            if (row_idx, col_idx) in merged_covered_set:
                col_idx += 1
                continue

            cell = cells[col_idx - 1]  # 因为 cells 是 0-based

            # 获取 rowspan/colspan（从预构建 map 中取）
            if (row_idx, col_idx) in merge_start_map:
                rowspan, colspan = merge_start_map[(row_idx, col_idx)]
            else:
                rowspan, colspan = 1, 1

            # 构建属性
            attrs = []
            if rowspan > 1:
                attrs.append(f'rowspan="{rowspan}"')
            if colspan > 1:
                attrs.append(f'colspan="{colspan}"')
            attr_str = " " + " ".join(attrs) if attrs else ""

            # 获取值
            if cell.value:
                value = str(cell.value)
            elif (row_idx, col_idx) in img_map:  # 图片处理
                value = get_image_type(img_map[(row_idx, col_idx)])
                print(
                    "cell image:", (row_idx, col_idx), len(img_map[(row_idx, col_idx)])
                )
            else:
                value = ""

            if cell.font.bold:  # xlsx中的粗体单元格视为HTML中的表头
                html.append(f"<th{attr_str}>{text_process(value)}</th>")
            else:
                html.append(f"<td{attr_str}>{text_process(value)}</td>")

            # 跳过跨列
            col_idx += colspan
        html.append("</tr>")
    html.append("</table>\n")

    # 只有<table></table>的空表
    if len(html) == 2:
        return ""
    return "".join(html)


def xlsx_to_html(xlsx_path):
    """
    将Excel表格转为HTML格式
    """
    full_html = StringIO()
    wb = load_workbook(xlsx_path, data_only=True)
    for idx, sheet in enumerate(wb.worksheets):
        # 转成HTML格式表格
        html_cnt = sheet_to_html(sheet.title, sheet)
        full_html.write(html_cnt)
    return full_html.getvalue()
